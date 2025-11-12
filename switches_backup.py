import tkinter as tk
from tkinter import filedialog, messagebox
import pandas as pd
import customtkinter as ctk
import threading
import os
import re
from rapidfuzz import fuzz, process
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

def clean_text(text):
    if isinstance(text, str):
        # Make lowercase and remove all non-alphanumeric characters (spaces, hyphens, etc.)
        cleaned = re.sub(r'[^a-z0-9]', '', text.lower())
        # Replace all 'o' with '0' to handle O/0 confusion
        cleaned = cleaned.replace('o', '0')
        return cleaned
    return text

# LoadingWindow class (copied from your other apps for consistency)
class LoadingWindow:
    """A modal window that displays processing progress."""
    def __init__(self, parent):
        self.window = ctk.CTkToplevel(parent)
        self.window.title("Processing")
        self.window.geometry("400x200")
        self.window.transient(parent)
        self.window.grab_set()
        # Center the window
        x = parent.winfo_x() + (parent.winfo_width() - 400) // 2
        y = parent.winfo_y() + (parent.winfo_height() - 200) // 2
        self.window.geometry(f"+{x}+{y}")
        self._create_widgets()
    def _create_widgets(self):
        self.frame = ctk.CTkFrame(self.window, fg_color="white")
        self.frame.pack(fill="both", expand=True, padx=20, pady=20)
        self.loading_label = ctk.CTkLabel(
            self.frame,
            text="Processing files...",
            font=("Segoe UI", 16, "bold"),
            text_color="#2c3e50"
        )
        self.loading_label.pack(pady=(20, 10))
        self.progress = ctk.CTkProgressBar(self.frame)
        self.progress.pack(fill="x", padx=20, pady=10)
        self.progress.set(0)
        self.status_label = ctk.CTkLabel(
            self.frame,
            text="Initializing...",
            font=("Segoe UI", 12),
            text_color="#7f8c8d"
        )
        self.status_label.pack(pady=10)
        self.window.protocol("WM_DELETE_WINDOW", lambda: None)
    def update_progress(self, value, status_text):
        self.progress.set(value)
        self.status_label.configure(text=status_text)
        self.window.update()
    def close(self):
        try:
            self.window.destroy()
        except tk.TclError:
            # This can happen if the window is already being destroyed,
            # which is fine.
            pass

class SwitchExtractorApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Switch Data Extractor")
        self.root.geometry("800x500")
        ctk.set_appearance_mode("light")
        ctk.set_default_color_theme("blue")
        self.root.state('zoomed')  # Start maximized (Windows)
        self.root.resizable(True, True)  # Allow resizing

        self.input_file_path = None
        self.distributor_files = []  # List of file paths for current month impalment
        self.extracted_df = None
        self.brokrage_file_path = None
        self.brokrage_file_paths = []
        self.scheme_master_path = None
        self.rows_to_highlight = []
        self.impalment_prev_files = []  # List of file paths for previous month impalment
        # self.rate_structure_file_paths = []  # Store multiple file paths
        # self.rate_structure_file_path = None

        # Main frame
        self.main_frame = ctk.CTkFrame(self.root, fg_color="white")
        self.main_frame.pack(fill="both", expand=True, padx=30, pady=30)

        # Title
        self.title_label = ctk.CTkLabel(
            self.main_frame,
            text="Switch Data Extractor",
            font=("Segoe UI", 28, "bold"),
            text_color="#2c3e50"
        )
        self.title_label.pack(pady=(0, 30))

        # Upload section (now using grid layout for better visibility)
        self.upload_frame = ctk.CTkFrame(self.main_frame, fg_color="#f8f9fa")
        self.upload_frame.pack(fill="x", padx=30, pady=10)

        # Row 0
        self.upload_label = ctk.CTkLabel(self.upload_frame, text="Upload Input File", font=("Segoe UI", 16, "bold"), text_color="#2c3e50")
        self.upload_label.grid(row=0, column=0, padx=10, pady=10, sticky="w")
        self.upload_btn = ctk.CTkButton(self.upload_frame, text="Upload Input File", command=self.upload_input_file, fg_color="#3498db", hover_color="#2980b9", font=("Segoe UI", 12), width=140, height=35)
        self.upload_btn.grid(row=0, column=1, padx=5, pady=10)
        self.input_file_label = ctk.CTkLabel(self.upload_frame, text="No file selected", text_color="#7f8c8d", font=("Segoe UI", 12))
        self.input_file_label.grid(row=0, column=2, padx=5, pady=10)

        self.upload_dist_btn = ctk.CTkButton(self.upload_frame, text="Upload Impalment Current Month File(s)", command=self.upload_distributor_files, fg_color="#8e44ad", hover_color="#6c3483", font=("Segoe UI", 12), width=260, height=35)
        self.upload_dist_btn.grid(row=0, column=3, padx=5, pady=10)
        self.dist_file_label = ctk.CTkLabel(self.upload_frame, text="No files selected", text_color="#7f8c8d", font=("Segoe UI", 12))
        self.dist_file_label.grid(row=0, column=4, padx=5, pady=10)

        # Row 1
        self.upload_prev_dist_btn = ctk.CTkButton(self.upload_frame, text="Upload Impalment Previous Month File(s)", command=self.upload_impalment_prev_files, fg_color="#6c3483", hover_color="#512e5f", font=("Segoe UI", 12), width=280, height=35)
        self.upload_prev_dist_btn.grid(row=1, column=0, padx=5, pady=10)
        self.prev_dist_file_label = ctk.CTkLabel(self.upload_frame, text="No files selected", text_color="#7f8c8d", font=("Segoe UI", 12))
        self.prev_dist_file_label.grid(row=1, column=1, padx=5, pady=10)

        self.upload_funding_btn = ctk.CTkButton(self.upload_frame, text="Distributor Funding File Payout", command=self.upload_funding_files, fg_color="#e67e22", hover_color="#ca6f1e", font=("Segoe UI", 12), width=260, height=35)
        self.upload_funding_btn.grid(row=1, column=2, padx=5, pady=10)
        self.funding_files_label = ctk.CTkLabel(self.upload_frame, text="No files selected", text_color="#7f8c8d", font=("Segoe UI", 12))
        self.funding_files_label.grid(row=1, column=3, padx=5, pady=10)

        # Add a Listbox to display uploaded funding files by month
        self.funding_files_listbox = tk.Listbox(self.main_frame, height=4, font=("Segoe UI", 12))
        self.funding_files_listbox.pack(fill="x", padx=40, pady=(0, 10))
        self.funding_files_listbox.insert(0, "No funding files uploaded")

        # Make columns expand to fill space
        for i in range(5):
            self.upload_frame.columnconfigure(i, weight=1)

        # --- Brokerage Structure Upload Section ---
        self.brokerage_frame = ctk.CTkFrame(self.main_frame, fg_color="#f8f9fa")
        self.brokerage_frame.pack(fill="x", padx=30, pady=10)

        brokerage_label = ctk.CTkLabel(
            self.brokerage_frame,
            text="Brokerage Structure Files",
            font=("Segoe UI", 18, "bold"),
            text_color="#2c3e50"
        )
        brokerage_label.pack(pady=(15, 10))

        brokerage_btn_frame = ctk.CTkFrame(self.brokerage_frame, fg_color="transparent")
        brokerage_btn_frame.pack(pady=5)

        self.upload_brokerage_multi_btn = ctk.CTkButton(
            brokerage_btn_frame,
            text="Upload Brokerage Structure Files",
            command=self.upload_brokerage_files_gui,
            fg_color="#f39c12",
            hover_color="#b9770e",
            font=("Segoe UI", 12),
            width=240,
            height=35
        )
        self.upload_brokerage_multi_btn.pack(side="left", padx=5)

        self.clear_brokerage_btn = ctk.CTkButton(
            brokerage_btn_frame,
            text="Clear Files",
            command=self.clear_brokerage_files,
            fg_color="#e74c3c",
            hover_color="#c0392b",
            font=("Segoe UI", 12),
            width=140,
            height=35
        )
        self.clear_brokerage_btn.pack(side="left", padx=5)

        self.brokerage_files_listbox = tk.Listbox(self.brokerage_frame, height=3, font=("Segoe UI", 12))
        self.brokerage_files_listbox.pack(fill="x", padx=10, pady=5)

        # --- Scheme Master File Upload Section ---
        self.scheme_master_frame = ctk.CTkFrame(self.main_frame, fg_color="#f8f9fa")
        self.scheme_master_frame.pack(fill="x", padx=30, pady=10)
        
        scheme_master_btn_frame = ctk.CTkFrame(self.scheme_master_frame, fg_color="transparent")
        scheme_master_btn_frame.pack(pady=10, padx=10, fill="x")

        ctk.CTkLabel(
            scheme_master_btn_frame,
            text="Scheme Master File",
            font=("Segoe UI", 16, "bold"),
            text_color="#2c3e50"
        ).pack(side="left", padx=(0, 20))

        self.upload_scheme_master_btn = ctk.CTkButton(
            scheme_master_btn_frame,
            text="Upload File",
            command=self.upload_scheme_master_file,
            fg_color="#16a085",
            hover_color="#1abc9c",
            font=("Segoe UI", 12)
        )
        self.upload_scheme_master_btn.pack(side="left", padx=10)

        self.scheme_master_file_label = ctk.CTkLabel(
            scheme_master_btn_frame,
            text="No file selected",
            font=("Segoe UI", 12),
            text_color="#7f8c8d"
        )
        self.scheme_master_file_label.pack(side="left", padx=10, fill="x")

        # --- Highlight Section ---
        self.highlight_frame = ctk.CTkFrame(self.main_frame, fg_color="#f8f9fa")
        self.highlight_frame.pack(fill="x", padx=30, pady=10)

        ctk.CTkLabel(
            self.highlight_frame,
            text="Highlight Specific Switches",
            font=("Segoe UI", 16, "bold"),
            text_color="#2c3e50"
        ).pack(pady=(10, 5))

        highlight_grid = ctk.CTkFrame(self.highlight_frame, fg_color="transparent")
        highlight_grid.pack(pady=5, padx=10, fill="x")
        highlight_grid.columnconfigure(1, weight=1)
        highlight_grid.columnconfigure(3, weight=1)

        ctk.CTkLabel(highlight_grid, text="Switch In Contains:", font=("Segoe UI", 12)).grid(row=0, column=0, padx=5, pady=5, sticky="w")
        self.highlight_in_entry = ctk.CTkEntry(highlight_grid, font=("Segoe UI", 12))
        self.highlight_in_entry.grid(row=0, column=1, padx=5, pady=5, sticky="ew")

        ctk.CTkLabel(highlight_grid, text="Switch Out Contains:", font=("Segoe UI", 12)).grid(row=0, column=2, padx=5, pady=5, sticky="w")
        self.highlight_out_entry = ctk.CTkEntry(highlight_grid, font=("Segoe UI", 12))
        self.highlight_out_entry.grid(row=0, column=3, padx=5, pady=5, sticky="ew")

        # Extract button
        self.extract_btn = ctk.CTkButton(
            self.main_frame, text="Extract Columns", command=self.start_extraction,
            fg_color="#27ae60", hover_color="#219a52", font=("Segoe UI", 14, "bold"), width=200, height=45
        )
        self.extract_btn.pack(pady=20)

        # Results area
        self.results_frame = ctk.CTkFrame(self.main_frame, fg_color="#f8f9fa")
        self.results_frame.pack(fill="both", expand=True, padx=10, pady=10)
        self.results_text = tk.Text(self.results_frame, height=10, width=80, font=("Consolas", 11), wrap="none")
        self.results_text.pack(side="left", fill="both", expand=True)
        self.scrollbar = ctk.CTkScrollbar(self.results_frame, command=self.results_text.yview)
        self.scrollbar.pack(side="right", fill="y")
        self.results_text.config(yscrollcommand=self.scrollbar.set)

        # Status bar
        self.status_frame = ctk.CTkFrame(self.main_frame, fg_color="#f8f9fa")
        self.status_frame.pack(fill="x", padx=10, pady=(0, 10))
        self.status_label = ctk.CTkLabel(
            self.status_frame,
            text="Status: Ready to upload files",
            font=("Segoe UI", 12),
            text_color="#2c3e50"
        )
        self.status_label.pack(pady=5)

    def upload_input_file(self):
        file_path = filedialog.askopenfilename(
            title="Select Input File",
            filetypes=[("Excel files", "*.xlsx *.xls"), ("CSV files", "*.csv"), ("All files", "*.*")]
        )
        if file_path:
            self.input_file_path = file_path
            self.input_file_label.configure(text=f"Selected: {os.path.basename(file_path)}", text_color="#27ae60")
            self.status_label.configure(text="Status: Input file uploaded!", text_color="#27ae60")

    def upload_distributor_files(self):
        file_paths = filedialog.askopenfilenames(
            title="Select Impalment Current Month File(s)",
            filetypes=[("Excel files", "*.xlsx *.xls"), ("CSV files", "*.csv"), ("All files", "*.*")]
        )
        if file_paths:
            self.distributor_files = list(file_paths)
            display_names = ", ".join([os.path.basename(f) for f in self.distributor_files])
            self.dist_file_label.configure(text=f"Selected: {display_names}", text_color="#27ae60")
            self.status_label.configure(text="Status: Distributor files uploaded!", text_color="#27ae60")
        else:
            self.dist_file_label.configure(text="No files selected", text_color="#7f8c8d")
            self.status_label.configure(text="Status: No distributor files uploaded!", text_color="#e74c3c")

    def upload_impalment_prev_files(self):
        file_paths = filedialog.askopenfilenames(
            title="Select Impalment Previous Month File(s)",
            filetypes=[("Excel files", "*.xlsx *.xls"), ("CSV files", "*.csv"), ("All files", "*.*")]
        )
        if file_paths:
            self.impalment_prev_files = list(file_paths)
            display_names = ", ".join([os.path.basename(f) for f in self.impalment_prev_files])
            self.prev_dist_file_label.configure(text=f"Selected: {display_names}", text_color="#27ae60")
            self.status_label.configure(text="Status: Previous month impalment files uploaded!", text_color="#27ae60")
        else:
            self.prev_dist_file_label.configure(text="No files selected", text_color="#7f8c8d")
            self.status_label.configure(text="Status: No previous month impalment files uploaded!", text_color="#e74c3c")

    def upload_brokrage_file(self):
        file_path = filedialog.askopenfilename(
            title="Select Brokrage Structure File",
            filetypes=[("Excel files", "*.xlsx *.xls"), ("All files", "*.*")]
        )
        if file_path:
            self.brokrage_file_path = file_path
            self.brokrage_file_label.configure(text=f"Selected: {os.path.basename(file_path)}", text_color="#27ae60")
            self.status_label.configure(text="Status: Brokrage structure file uploaded!", text_color="#27ae60")

    def upload_brokrage_files(self):
        file_paths = filedialog.askopenfilenames(
            title="Select Brokrage Structure Files (Multiple)",
            filetypes=[("Excel files", "*.xlsx *.xls"), ("All files", "*.*")]
        )
        if file_paths:
            self.brokrage_file_paths = list(file_paths)
            display_names = ", ".join([os.path.basename(fp) for fp in self.brokrage_file_paths])
            self.brokrage_files_label.configure(text=f"Selected: {display_names}", text_color="#27ae60")
            self.status_label.configure(text="Status: Multiple brokrage structure files uploaded!", text_color="#27ae60")

    def upload_brokerage_files_gui(self):
        file_paths = filedialog.askopenfilenames(
            title="Select Brokerage Structure Files",
            filetypes=[("Excel files", "*.xlsx *.xls"), ("All files", "*.*")]
        )
        if file_paths:
            self.brokrage_file_paths = list(file_paths)
            self.brokerage_files_listbox.delete(0, tk.END)
            for fp in self.brokrage_file_paths:
                self.brokerage_files_listbox.insert(tk.END, os.path.basename(fp))
            self.status_label.configure(text=f"Status: Uploaded {len(self.brokrage_file_paths)} brokerage structure files!", text_color="#27ae60")

    def clear_brokerage_files(self):
        self.brokrage_file_paths = []
        self.brokerage_files_listbox.delete(0, tk.END)
        self.status_label.configure(text="Status: Cleared all brokerage structure files", text_color="#e74c3c")

    def upload_scheme_master_file(self):
        file_path = filedialog.askopenfilename(
            title="Select Scheme Master File",
            filetypes=[("Excel files", "*.xlsx *.xls"), ("All files", "*.*")]
        )
        if file_path:
            self.scheme_master_path = file_path
            self.scheme_master_file_label.configure(text=f"Selected: {os.path.basename(file_path)}", text_color="#27ae60")
            self.status_label.configure(text="Status: Scheme master file uploaded!", text_color="#16a085")

    def upload_funding_files(self):
        import re
        file_paths = filedialog.askopenfilenames(
            title="Select Distributor Funding File Payout(s)",
            filetypes=[("Excel files", "*.xlsx *.xls"), ("CSV files", "*.csv"), ("All files", "*.*")]
        )
        if file_paths:
            self.funding_files = list(file_paths)
            display_names = ", ".join([os.path.basename(f) for f in self.funding_files])
            self.funding_files_label.configure(text=f"Selected: {display_names}", text_color="#27ae60")
            self.status_label.configure(text="Status: Funding files uploaded!", text_color="#27ae60")
            # Update the Listbox with month and file name
            self.funding_files_listbox.delete(0, tk.END)
            month_pattern = re.compile(r"FundingSummary_([A-Za-z]+\d{4})", re.IGNORECASE)
            for f in self.funding_files:
                fname = os.path.basename(f)
                match = month_pattern.search(fname)
                if match:
                    month = match.group(1)
                    self.funding_files_listbox.insert(tk.END, f"{month}: {fname}")
                else:
                    self.funding_files_listbox.insert(tk.END, fname)
        else:
            self.funding_files_label.configure(text="No files selected", text_color="#7f8c8d")
            self.status_label.configure(text="Status: No funding files uploaded!", text_color="#e74c3c")
            self.funding_files_listbox.delete(0, tk.END)
            self.funding_files_listbox.insert(0, "No funding files uploaded")

    def start_extraction(self):
        if not self.input_file_path:
            messagebox.showerror("Error", "Please upload the input file first!")
            return
        self.status_label.configure(text="Status: Extracting columns...", text_color="#2980b9")
        self.loading_window = LoadingWindow(self.root)
        thread = threading.Thread(target=self.extract_columns)
        thread.start()

    def extract_columns(self):
        try:
            self.loading_window.update_progress(0.05, "Reading input file...")
            # Read the input file
            if self.input_file_path.endswith('.csv'):
                df = pd.read_csv(self.input_file_path)
            else:
                df = pd.read_excel(self.input_file_path)
            self.loading_window.update_progress(0.15, "Processing columns...")
            # Extract columns (case-insensitive)
            required_cols = ['SWITCH_DETAILS_FOLIO_NO', 'TRADES_BROK_DLR_CODE', 'TRADES_AMOUNT', 'LONG_NAME1', 'LONG_NAME']
            df.columns = [col.upper() for col in df.columns]
            # Remove rows where TRADES_BROK_DLR_CODE is '000000-0'
            if 'TRADES_BROK_DLR_CODE' in df.columns:
                df = df[df['TRADES_BROK_DLR_CODE'] != '000000-0']
            extracted = df[[col for col in required_cols if col in df.columns]]

            # If distributor file is provided, merge RateCategory
            if self.distributor_files:
                self.loading_window.update_progress(0.25, "Merging distributor files...")
                for distributor_file in self.distributor_files:
                    if distributor_file.endswith('.csv'):
                        dist_df = pd.read_csv(distributor_file)
                    else:
                        dist_df = pd.read_excel(distributor_file)
                    dist_df.columns = [col.upper() for col in dist_df.columns]
                    if 'AGENT' in dist_df.columns and 'RATECATEGORY' in dist_df.columns:
                        extracted = extracted.merge(
                            dist_df[['AGENT', 'RATECATEGORY']],
                            left_on='TRADES_BROK_DLR_CODE',
                            right_on='AGENT',
                            how='left'
                        )
                        if 'AGENT' in extracted.columns:
                            extracted = extracted.drop(columns=['AGENT'])
            if self.impalment_prev_files:
                self.loading_window.update_progress(0.27, "Merging previous month impalment files...")
                for prev_file in self.impalment_prev_files:
                    if prev_file.endswith('.csv'):
                        prev_df = pd.read_csv(prev_file)
                    else:
                        prev_df = pd.read_excel(prev_file)
                    prev_df.columns = [col.upper() for col in prev_df.columns]
                    if 'AGENT' in prev_df.columns and 'RATECATEGORY' in prev_df.columns:
                        extracted = extracted.merge(
                            prev_df[['AGENT', 'RATECATEGORY']].rename(columns={'RATECATEGORY': 'RATECATEGORY -Previous'}),
                            left_on='TRADES_BROK_DLR_CODE',
                            right_on='AGENT',
                            how='left',
                            suffixes=(None, '_prev')
                        )
                        if 'AGENT' in extracted.columns:
                            extracted = extracted.drop(columns=['AGENT'])

            # --- Funding File Payout Merge Logic ---
            import re
            month_abbr = {
                'january': 'JAN', 'february': 'FEB', 'march': 'MAR', 'april': 'APR', 'may': 'MAY', 'june': 'JUN',
                'july': 'JUL', 'august': 'AUG', 'september': 'SEP', 'october': 'OCT', 'november': 'NOV', 'december': 'DEC'
            }
            if hasattr(self, 'funding_files') and self.funding_files:
                for funding_file in self.funding_files:
                    fname = os.path.basename(funding_file)
                    # Extract month from filename
                    m = re.search(r'FundingSummary_([A-Za-z]+)\d{4}', fname)
                    if m:
                        month_str = m.group(1).lower()
                        month_col = month_abbr.get(month_str, month_str.upper())
                        colname = f'Net_Amount {month_col}'
                    else:
                        colname = f'Net_Amount {fname}'
                    # Try to find header row (row 4, 5, or 6)
                    header_row = None
                    for i in range(3, 6):
                        try:
                            temp_df = pd.read_excel(funding_file, header=i)
                            if 'AgentCode' in temp_df.columns and 'Net_Amount' in temp_df.columns:
                                header_row = i
                                break
                        except Exception:
                            continue
                    if header_row is not None:
                        fund_df = pd.read_excel(funding_file, header=header_row)
                        # Only keep AgentCode and Net_Amount
                        fund_df = fund_df[['AgentCode', 'Net_Amount']]
                        fund_df = fund_df.rename(columns={'AgentCode': 'TRADES_BROK_DLR_CODE'})
                        # Merge into main extracted DataFrame
                        extracted = extracted.merge(
                            fund_df,
                            on='TRADES_BROK_DLR_CODE',
                            how='left',
                            suffixes=(None, f'_{month_col}')
                        )
                        # Rename Net_Amount column
                        if 'Net_Amount' in extracted.columns:
                            extracted = extracted.rename(columns={'Net_Amount': colname})
                    else:
                        print(f"Could not find header row in funding file: {fname}")
            # --- End Funding File Payout Merge Logic ---

            # Prepare Scheme Master lookup
            scheme_lookup = {}
            if self.scheme_master_path:
                self.loading_window.update_progress(0.30, "Reading scheme master...")
                print("\n--- Loading Scheme Master File ---")
                try:
                    scheme_df = pd.read_excel(self.scheme_master_path)
                    print(f"DEBUG | Successfully loaded scheme master. Found {len(scheme_df)} rows.")
                    print(f"DEBUG | Original Columns: {list(scheme_df.columns)}")
                    
                    # Normalize column names for robust access
                    scheme_df.columns = [normalize_colname(col) for col in scheme_df.columns]
                    print(f"DEBUG | Normalized Columns: {list(scheme_df.columns)}")

                    if 'scheme' in scheme_df.columns and 'schemetype' in scheme_df.columns:
                        print("DEBUG | 'scheme' and 'schemetype' columns found.")
                        # Create a normalized version of the fund names for matching
                        scheme_df['NORMALIZED_SCHEME'] = scheme_df['scheme'].apply(
                            lambda x: normalize_fund_name(extract_core_fund_name(x))
                        )
                        # Create a lookup dictionary for fast matching
                        scheme_lookup = pd.Series(
                            scheme_df['schemetype'].values,
                            index=scheme_df['NORMALIZED_SCHEME']
                        ).to_dict()
                        print(f"DEBUG | Created lookup dictionary with {len(scheme_lookup)} entries.")
                        # Print first 5 items for verification
                        print("DEBUG | Sample of lookup dictionary:")
                        for i, (k, v) in enumerate(scheme_lookup.items()):
                            if i >= 5: break
                            print(f"  '{k}' -> '{v}'")
                    else:
                        print("ERROR | 'scheme' or 'schemetype' column not found in Scheme Master!")

                except Exception as e:
                    print(f"ERROR | Could not process Scheme Master file. Error: {e}")
                print("--- Finished Loading Scheme Master ---\n")

            self.loading_window.update_progress(0.35, "Preparing for matching...")
            col_order = ['TRADES_BROK_DLR_CODE', 'RATECATEGORY']
            rest_cols = [col for col in extracted.columns if col not in col_order]
            extracted = extracted[[col for col in col_order if col in extracted.columns] + rest_cols]

            # --- Start of new Brokerage Matching Logic ---
            brokerage_sheets = {}
            # Pre-load all sheets from all brokerage files
            if self.brokrage_file_paths:
                self.loading_window.update_progress(0.40, "Loading brokerage files...")
                for file_path in self.brokrage_file_paths:
                    try:
                        xls = pd.ExcelFile(file_path)
                        for sheet_name in xls.sheet_names:
                            cleaned_sheet_name = clean_text(sheet_name)
                            if cleaned_sheet_name not in brokerage_sheets:
                                brokerage_sheets[cleaned_sheet_name] = pd.read_excel(xls, sheet_name=sheet_name)
                    except Exception as e:
                        print(f"Error reading and caching file {file_path}: {e}")
            elif self.brokrage_file_path:
                self.loading_window.update_progress(0.40, "Loading brokerage file...")
                try:
                    xls = pd.ExcelFile(self.brokrage_file_path)
                    for sheet_name in xls.sheet_names:
                        cleaned_sheet_name = clean_text(sheet_name)
                        if cleaned_sheet_name not in brokerage_sheets:
                            brokerage_sheets[cleaned_sheet_name] = pd.read_excel(xls, sheet_name=sheet_name)
                except Exception as e:
                    print(f"Error reading and caching file {self.brokrage_file_path}: {e}")

            switch_in_trails = []
            switch_out_trails = []
            switch_in_status = []
            switch_out_status = []
            switch_in_scheme_types = []
            switch_out_scheme_types = []
            switch_in_trails_prev = []
            switch_in_status_prev = []

            # Process each row against the pre-loaded sheets
            for idx, row in extracted.iterrows():
                if (idx + 1) % 100 == 0:
                    progress = 0.5 + 0.4 * (idx + 1) / len(extracted)
                    self.loading_window.update_progress(progress, f"Matching row {idx + 1}/{len(extracted)}...")

                rate_category_raw = row.get('RATECATEGORY')
                
                # Match for switch in (LONG_NAME)
                switch_in_fund = row.get('LONG_NAME')
                in_trail, in_found = self._find_trail_value(switch_in_fund, rate_category_raw, brokerage_sheets)
                switch_in_trails.append(in_trail)
                switch_in_status.append('Found' if in_found else 'Not Found')
                
                # Previous switch in trail using RATECATEGORY -Previous

                rate_category_prev = row.get('RATECATEGORY -Previous')
                in_trail_prev, in_found_prev = self._find_trail_value(switch_in_fund, rate_category_prev, brokerage_sheets)
                switch_in_trails_prev.append(in_trail_prev)
                switch_in_status_prev.append('Found' if in_found_prev else 'Not Found')

                # Get scheme type for switch in
                in_scheme_type = self._get_scheme_type(switch_in_fund, scheme_lookup, idx)
                switch_in_scheme_types.append(in_scheme_type)

                # Match for switch out (LONG_NAME1)
                switch_out_fund = row.get('LONG_NAME1')
                out_trail, out_found = self._find_trail_value(switch_out_fund, rate_category_raw, brokerage_sheets)
                switch_out_trails.append(out_trail)
                switch_out_status.append('Found' if out_found else 'Not Found')
                
                # Get scheme type for switch out
                out_scheme_type = self._get_scheme_type(switch_out_fund, scheme_lookup, idx)
                switch_out_scheme_types.append(out_scheme_type)

            extracted['Scheme Type Swith IN'] = switch_in_scheme_types
            extracted['Scheme Type Swith Out'] = switch_out_scheme_types
            extracted['switch in TRAIL_1ST_YEAR'] = switch_in_trails
            extracted['switch in MATCH_STATUS'] = switch_in_status
            extracted['switch in TRAIL_1ST_YEAR -Previous'] = switch_in_trails_prev
            extracted['switch in MATCH_STATUS -Previous'] = switch_in_status_prev
            extracted['switch out TRAIL_1ST_YEAR'] = switch_out_trails
            extracted['switch out MATCH_STATUS'] = switch_out_status

            # Add the new column: previous < current switch in TRAIL_1ST_YEAR
            prev_trails = pd.to_numeric(extracted['switch in TRAIL_1ST_YEAR -Previous'], errors='coerce')
            curr_trails = pd.to_numeric(extracted['switch in TRAIL_1ST_YEAR'], errors='coerce')
            extracted['previous < current switch in TRAIL_1ST_YEAR'] = ''
            extracted.loc[(curr_trails > prev_trails).fillna(False), 'previous < current switch in TRAIL_1ST_YEAR'] = 'check'

            # Add the switching rate check column
            in_trails = pd.to_numeric(extracted['switch in TRAIL_1ST_YEAR'], errors='coerce')
            out_trails = pd.to_numeric(extracted['switch out TRAIL_1ST_YEAR'], errors='coerce')
            
            extracted['switching rate check'] = ''
            extracted.loc[(in_trails > out_trails).fillna(False), 'switching rate check'] = 'check'

            # --- New 'Direct to Regular' Logic ---
            
            # Condition 1: Check if both scheme types are 'Equity Funds'
            is_equity_in = extracted['Scheme Type Swith IN'].str.strip().str.lower() == 'equity funds'
            is_equity_out = extracted['Scheme Type Swith Out'].str.strip().str.lower() == 'equity funds'
            
            # Condition 2: Check for a switch from a 'Direct' plan to a 'Regular' plan
            switch_in_name_raw = extracted['LONG_NAME'].astype(str).str.lower()
            switch_out_name_raw = extracted['LONG_NAME1'].astype(str).str.lower()
            is_regular_in = switch_in_name_raw.str.contains('regular', na=False)
            is_direct_out = switch_out_name_raw.str.contains('direct', na=False)
            
            # Condition 3: Check if the core fund name is the same for both
            core_in_name = extracted['LONG_NAME'].apply(lambda x: normalize_fund_name(extract_core_fund_name(x)))
            core_out_name = extracted['LONG_NAME1'].apply(lambda x: normalize_fund_name(extract_core_fund_name(x)))
            is_same_core_name = core_in_name == core_out_name
            
            # Combine all conditions
            final_condition = (
                is_equity_in &
                is_equity_out &
                is_regular_in &
                is_direct_out &
                is_same_core_name
            )
            
            extracted['Direct to Regular'] = ''
            extracted.loc[final_condition.fillna(False), 'Direct to Regular'] = 'check'
            
            # --- End of new 'Direct to Regular' Logic ---

            # --- End of new Brokerage Matching Logic ---

            self.loading_window.update_progress(0.95, "Saving extracted data...")
            # Rename columns for output
            rename_map = {
                'LONG_NAME': 'switch in',
                'LONG_NAME1': 'switch out'
            }
            extracted = extracted.rename(columns=rename_map)
            
            # Define final column order and remove unwanted columns
            final_col_order = [
                'SWITCH_DETAILS_FOLIO_NO', 'TRADES_BROK_DLR_CODE', 'RATECATEGORY', 'RATECATEGORY -Previous', 'Effective Date', 'TRADES_AMOUNT',
                'Scheme Type Swith IN', 'switch in', 'switch out', 'Scheme Type Swith Out',
                'switch in TRAIL_1ST_YEAR', 'switch in TRAIL_1ST_YEAR -Previous',
                'previous < current switch in TRAIL_1ST_YEAR',
                'switch out TRAIL_1ST_YEAR',
                'switching rate check', 'Direct to Regular'
            ]
            # Add all Net_Amount columns (for each month) to the output, after the standard columns
            net_amount_cols = [col for col in extracted.columns if col.startswith('Net_Amount') and col not in final_col_order]
            # Rename Net_Amount columns to PAYOUT <MONTH>
            payout_col_map = {col: col.replace('Net_Amount', 'PAYOUT') for col in net_amount_cols}
            extracted = extracted.rename(columns=payout_col_map)
            payout_cols = [payout_col_map[col] for col in net_amount_cols]
            final_col_order += payout_cols
            
            # --- Highlighting Logic (without adding a column) ---
            highlight_in_text = self.highlight_in_entry.get().strip().lower()
            highlight_out_text = self.highlight_out_entry.get().strip().lower()
            self.rows_to_highlight = []  # Reset before each extraction

            if highlight_in_text and highlight_out_text:
                print("\n--- Highlighting Logic ---")
                print(f"DEBUG | Highlight 'Switch In' contains: '{highlight_in_text}'")
                print(f"DEBUG | Highlight 'Switch Out' contains: '{highlight_out_text}'")

                if 'switch in' in extracted.columns and 'switch out' in extracted.columns:
                    match_in = extracted['switch in'].str.lower().str.contains(highlight_in_text, na=False)
                    match_out = extracted['switch out'].str.lower().str.contains(highlight_out_text, na=False)
                    
                    print(f"DEBUG | Found {match_in.sum()} potential 'in' matches.")
                    print(f"DEBUG | Found {match_out.sum()} potential 'out' matches.")

                    # Store the indices of rows to be highlighted
                    self.rows_to_highlight = extracted[match_in & match_out].index.tolist()
                    print(f"DEBUG | Final rows to highlight (by index): {self.rows_to_highlight}")
                else:
                    print("ERROR | 'switch in' or 'switch out' columns not found for highlighting.")
                print("--- End Highlighting Logic ---\n")

            # Filter to only include columns that actually exist in the dataframe
            final_cols_to_keep = [col for col in final_col_order if col in extracted.columns]
            extracted = extracted[final_cols_to_keep]

            self.extracted_df = extracted
            
            # Format output professionally
            self.display_professional_results(extracted)
            
            self.status_label.configure(text="Status: Extraction complete!", text_color="#27ae60")
            self.loading_window.close()
            self.save_extracted()
        except Exception as e:
            self.loading_window.close()
            self.results_text.delete(1.0, tk.END)
            
            # Format error message professionally
            error_header = "╔" + "═" * 78 + "╗\n"
            error_header += "║" + "EXTRACTION ERROR".center(78) + "║\n"
            error_header += "╚" + "═" * 78 + "╝\n\n"
            
            error_details = "┌" + "─" * 58 + "┐\n"
            error_details += "│ ERROR DETAILS".ljust(58) + "│\n"
            error_details += "├" + "─" * 58 + "┤\n"
            error_details += f"│ Error Type: {type(e).__name__}".ljust(58) + "│\n"
            error_details += f"│ Error Message: {str(e)}".ljust(58) + "│\n"
            error_details += f"│ Timestamp: {pd.Timestamp.now().strftime('%Y-%m-%d %H:%M:%S')}".ljust(58) + "│\n"
            error_details += "└" + "─" * 58 + "┘\n\n"
            
            error_suggestions = "┌" + "─" * 58 + "┐\n"
            error_suggestions += "│ TROUBLESHOOTING SUGGESTIONS".ljust(58) + "│\n"
            error_suggestions += "├" + "─" * 58 + "┤\n"
            error_suggestions += "│ 1. Ensure all required files are uploaded".ljust(58) + "│\n"
            error_suggestions += "│ 2. Check that file formats are supported".ljust(58) + "│\n"
            error_suggestions += "│ 3. Verify that input files contain required columns".ljust(58) + "│\n"
            error_suggestions += "│ 4. Ensure brokerage files have proper sheet structure".ljust(58) + "│\n"
            error_suggestions += "│ 5. Check file permissions and accessibility".ljust(58) + "│\n"
            error_suggestions += "└" + "─" * 58 + "┘\n\n"
            
            error_footer = "╔" + "═" * 78 + "╗\n"
            error_footer += "║" + "Please review the error and try again".center(78) + "║\n"
            error_footer += "╚" + "═" * 78 + "╝\n"
            
            full_error = error_header + error_details + error_suggestions + error_footer
            self.results_text.insert(tk.END, full_error)
            
            # Apply formatting to error display with enhanced styling
            self.results_text.tag_configure("error_main_header", font=("Consolas", 14, "bold"), foreground="#e74c3c")
            self.results_text.tag_configure("error_section_header", font=("Consolas", 12, "bold"), foreground="#c0392b")
            self.results_text.tag_configure("error_border", font=("Consolas", 10), foreground="#95a5a6")
            self.results_text.tag_configure("error_data", font=("Consolas", 10), foreground="#2c3e50")
            self.results_text.tag_configure("error_final", font=("Consolas", 12, "bold"), foreground="#e74c3c")
            
            # Apply tags with dynamic line detection
            lines = full_error.split('\n')
            current_line = 1
            
            for line in lines:
                if line.strip().startswith('╔') or line.strip().startswith('╚'):
                    # Main border lines
                    start_pos = f"{current_line}.0"
                    end_pos = f"{current_line}.{len(line)}"
                    if "EXTRACTION ERROR" in line:
                        self.results_text.tag_add("error_main_header", start_pos, end_pos)
                    else:
                        self.results_text.tag_add("error_final", start_pos, end_pos)
                elif line.strip().startswith('│') and not line.strip().startswith('├') and not line.strip().startswith('└'):
                    # Section headers
                    start_pos = f"{current_line}.0"
                    end_pos = f"{current_line}.{len(line)}"
                    self.results_text.tag_add("error_section_header", start_pos, end_pos)
                elif line.strip().startswith('┌') or line.strip().startswith('├') or line.strip().startswith('└'):
                    # Border lines
                    start_pos = f"{current_line}.0"
                    end_pos = f"{current_line}.{len(line)}"
                    self.results_text.tag_add("error_border", start_pos, end_pos)
                current_line += 1
            
            self.status_label.configure(text="Status: Extraction failed!", text_color="#e74c3c")

    def _get_scheme_type(self, fund_name, scheme_lookup, row_index):
        """Find scheme type using direct and fuzzy matching with debug output."""
        if pd.isna(fund_name) or not scheme_lookup:
            return ''
            
        normalized_fund = normalize_fund_name(extract_core_fund_name(fund_name))
        
        # --- Start of Debug Prints ---
        # Only print for the first 5 rows to avoid spamming the console
        if row_index < 5:
            print("-" * 50)
            print(f"DEBUG | Row {row_index} | Original Fund Name: '{fund_name}'")
            print(f"DEBUG | Row {row_index} | Cleaned & Normalized: '{normalized_fund}'")
        # --- End of Debug Prints ---

        # 1. Try direct lookup with normalization for speed
        scheme_type = scheme_lookup.get(normalized_fund)
        if scheme_type:
            if row_index < 5:
                print(f"DEBUG | Row {row_index} | Match Type: Direct Match Found!")
            return scheme_type
            
        # 2. If direct lookup fails, use fuzzy matching as a fallback
        best_match = process.extractOne(
            normalized_fund,
            scheme_lookup.keys(),
            scorer=fuzz.token_set_ratio,
            score_cutoff=85  # Lowered threshold to be less strict
        )
        
        # --- Start of Debug Prints ---
        if row_index < 5:
            if best_match:
                print(f"DEBUG | Row {row_index} | Match Type: Fuzzy Match Found!")
                print(f"DEBUG | Row {row_index} | Best Match: '{best_match[0]}' (Score: {best_match[1]:.1f})")
            else:
                print(f"DEBUG | Row {row_index} | Match Type: No Match Found (Score < 85)")
        # --- End of Debug Prints ---

        if best_match:
            return scheme_lookup[best_match[0]]
            
        return ''

    def _find_trail_value(self, fund_name, rate_category_raw, brokerage_sheets):
        """Helper function to find trail value for a given fund and rate category."""
        if pd.isna(fund_name) or pd.isna(rate_category_raw):
            return None, False

        rate_category = clean_text(rate_category_raw)
        matching_sheet_key = find_best_sheet(rate_category, brokerage_sheets, threshold=85)

        if matching_sheet_key:
            try:
                sheet_df = brokerage_sheets[matching_sheet_key]
                normalized_fund_name = normalize_fund_name(extract_core_fund_name(fund_name))
                
                required_brokerage_cols = ['Name of the Fund', 'Trail (% p.a.) 1st year']
                header_row = find_header_row(sheet_df, required_brokerage_cols)
                
                if header_row is not None:
                    processed_df = sheet_df.copy()
                    processed_df.columns = processed_df.iloc[header_row]
                    processed_df = processed_df.iloc[header_row + 1:]
                    processed_df.columns = [normalize_colname(col) for col in processed_df.columns]
                    
                    if 'nameofthefund' in processed_df.columns:
                        processed_df['NORMALIZED_FUND'] = processed_df['nameofthefund'].apply(lambda x: normalize_fund_name(extract_core_fund_name(x)))
                        match = processed_df[processed_df['NORMALIZED_FUND'] == normalized_fund_name]
                        
                        if not match.empty:
                            trail_col = next((col for col in processed_df.columns if 'trail' in col and '1st' in col and 'year' in col), None)
                            if trail_col:
                                trail = match.iloc[0][trail_col]
                                return trail, True
            except Exception as e:
                print(f"Error processing match for {fund_name} in '{matching_sheet_key}': {e}")
        
        return None, False

    def display_professional_results(self, df):
        """Display results in a professional format with summary statistics"""
        self.results_text.delete(1.0, tk.END)
        
        # Header with enhanced formatting
        header = "╔" + "═" * 78 + "╗\n"
        header += "║" + "SWITCH DATA EXTRACTION RESULTS".center(78) + "║\n"
        header += "╚" + "═" * 78 + "╝\n\n"
        
        # Summary Statistics with border
        summary = "┌" + "─" * 38 + "┐\n"
        summary += "│ SUMMARY STATISTICS".ljust(38) + "│\n"
        summary += "├" + "─" * 38 + "┤\n"
        summary += f"│ Total Records Processed: {len(df):,}".ljust(38) + "│\n"
        
        if 'SHEET_MATCH_STATUS' in df.columns:
            match_stats = df['SHEET_MATCH_STATUS'].value_counts()
            summary += f"│ Successful Matches: {match_stats.get('Found', 0):,}".ljust(38) + "│\n"
            summary += f"│ Unmatched Records: {match_stats.get('Not Found', 0):,}".ljust(38) + "│\n"
            match_rate = (match_stats.get('Found', 0) / len(df)) * 100 if len(df) > 0 else 0
            summary += f"│ Match Success Rate: {match_rate:.1f}%".ljust(38) + "│\n"
        
        if 'TRAIL_1ST_YEAR' in df.columns:
            trail_stats = df['TRAIL_1ST_YEAR'].notna().sum()
            summary += f"│ Records with Trail Data: {trail_stats:,}".ljust(38) + "│\n"
            summary += f"│ Records Missing Trail Data: {len(df) - trail_stats:,}".ljust(38) + "│\n"
        
        summary += "└" + "─" * 38 + "┘\n\n"
        
        # Column Information with border
        columns_info = "┌" + "─" * 58 + "┐\n"
        columns_info += "│ COLUMNS EXTRACTED".ljust(58) + "│\n"
        columns_info += "├" + "─" * 58 + "┤\n"
        for i, col in enumerate(df.columns, 1):
            non_null_count = df[col].notna().sum()
            null_count = df[col].isna().sum()
            col_line = f"│ {i:2d}. {col:<25} │ Non-null: {non_null_count:>6,} │ Null: {null_count:>6,} │"
            columns_info += col_line + "\n"
        columns_info += "└" + "─" * 58 + "┘\n\n"
        
        # Sample Data with enhanced border formatting
        sample_data = "┌" + "─" * 78 + "┐\n"
        sample_data += "│ SAMPLE DATA (First 10 Records)".ljust(78) + "│\n"
        sample_data += "├" + "─" * 78 + "┤\n"
        
        # Format the sample data with proper alignment and borders
        if len(df) > 0:
            sample_df = df.head(10)
            
            # Get column widths for proper formatting
            col_widths = {}
            for col in sample_df.columns:
                max_width = len(str(col))
                for val in sample_df[col]:
                    max_width = max(max_width, len(str(val)))
                col_widths[col] = min(max_width, 15)  # Cap at 15 characters for better fit
            
            # Calculate total width for border alignment
            total_width = sum(col_widths.values()) + len(col_widths) * 3 + 1  # 3 for separators, 1 for padding
            
            # Header row with borders
            header_row = "│ "
            for col in sample_df.columns:
                header_row += f"{col[:col_widths[col]]:<{col_widths[col]}} │ "
            header_row = header_row.ljust(78) + "│"
            sample_data += header_row + "\n"
            
            # Separator line
            separator = "├"
            for col in sample_df.columns:
                separator += "─" * (col_widths[col] + 2) + "┼"
            separator = separator[:-1] + "┤"  # Replace last ┼ with ┤
            separator = separator.ljust(78) + "│"
            sample_data += separator + "\n"
            
            # Data rows with borders
            for _, row in sample_df.iterrows():
                data_row = "│ "
                for col in sample_df.columns:
                    val = str(row[col])[:col_widths[col]] if pd.notna(row[col]) else "N/A"
                    data_row += f"{val:<{col_widths[col]}} │ "
                data_row = data_row.ljust(78) + "│"
                sample_data += data_row + "\n"
        else:
            sample_data += "│ No data available".ljust(78) + "│\n"
        
        sample_data += "└" + "─" * 78 + "┘\n\n"
        
        # Processing Information with border
        processing_info = "┌" + "─" * 58 + "┐\n"
        processing_info += "│ PROCESSING INFORMATION".ljust(58) + "│\n"
        processing_info += "├" + "─" * 58 + "┤\n"
        processing_info += f"│ Input File: {os.path.basename(self.input_file_path) if self.input_file_path else 'Not provided'}".ljust(58) + "│\n"
        processing_info += f"│ Distributor Files: {len(self.distributor_files)} files loaded".ljust(58) + "│\n"
        processing_info += f"│ Previous Month Impalment Files: {len(self.impalment_prev_files)} files loaded".ljust(58) + "│\n"
        processing_info += f"│ Processing Date: {pd.Timestamp.now().strftime('%Y-%m-%d %H:%M:%S')}".ljust(58) + "│\n"
        processing_info += "└" + "─" * 58 + "┘\n\n"
        
        # Final message with border
        final_msg = "╔" + "═" * 78 + "╗\n"
        final_msg += "║" + "EXTRACTION COMPLETE - Data ready for saving".center(78) + "║\n"
        final_msg += "╚" + "═" * 78 + "╝\n"
        
        # Combine all sections
        full_output = header + summary + columns_info + sample_data + processing_info + final_msg
        
        # Insert into text widget with proper formatting
        self.results_text.insert(tk.END, full_output)
        
        # Configure text widget for better readability with enhanced styling
        self.results_text.tag_configure("main_header", font=("Consolas", 14, "bold"), foreground="#2c3e50")
        self.results_text.tag_configure("section_header", font=("Consolas", 12, "bold"), foreground="#34495e")
        self.results_text.tag_configure("border", font=("Consolas", 10), foreground="#7f8c8d")
        self.results_text.tag_configure("data", font=("Consolas", 10), foreground="#2c3e50")
        self.results_text.tag_configure("final_header", font=("Consolas", 12, "bold"), foreground="#27ae60")
        
        # Apply tags for better formatting
        # Main header
        self.results_text.tag_add("main_header", "1.0", "4.0")
        
        # Section headers (lines with │ at start)
        lines = full_output.split('\n')
        current_line = 5  # Start after main header
        
        for line in lines[4:]:  # Skip main header lines
            if line.strip().startswith('│') and not line.strip().startswith('├') and not line.strip().startswith('└'):
                # This is a section header
                start_pos = f"{current_line}.0"
                end_pos = f"{current_line}.{len(line)}"
                self.results_text.tag_add("section_header", start_pos, end_pos)
            elif line.strip().startswith('╔') or line.strip().startswith('╚'):
                # This is a main border
                start_pos = f"{current_line}.0"
                end_pos = f"{current_line}.{len(line)}"
                if "EXTRACTION COMPLETE" in line:
                    self.results_text.tag_add("final_header", start_pos, end_pos)
                else:
                    self.results_text.tag_add("main_header", start_pos, end_pos)
            elif line.strip().startswith('┌') or line.strip().startswith('├') or line.strip().startswith('└'):
                # This is a border line
                start_pos = f"{current_line}.0"
                end_pos = f"{current_line}.{len(line)}"
                self.results_text.tag_add("border", start_pos, end_pos)
            current_line += 1
        
        # Scroll to top
        self.results_text.see("1.0")

    def save_extracted(self):
        if self.extracted_df is None:
            messagebox.showerror("Error", "No data to save!")
            return
        save_path = filedialog.asksaveasfilename(
            title="Save Extracted Data",
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
        )
        if save_path:
            try:
                # Create a professional Excel file with formatting
                self.create_formatted_excel(save_path)
                messagebox.showinfo("Success", f"Formatted data saved to: {save_path}")
                self.status_label.configure(text="Status: Formatted data saved!", text_color="#27ae60")
            except Exception as e:
                messagebox.showerror("Error", f"Could not save file: {e}")
                self.status_label.configure(text="Status: Save failed!", text_color="#e74c3c")

    def create_formatted_excel(self, file_path):
        """Create a professionally formatted Excel file"""
        try:
            # Try to import xlsxwriter
            import xlsxwriter
            self._create_formatted_excel_xlsxwriter(file_path)
        except ImportError:
            # Fallback to openpyxl if xlsxwriter is not available
            print("xlsxwriter not available, using openpyxl for formatting...")
            self._create_formatted_excel_openpyxl(file_path)
        except Exception as e:
            # Final fallback to simple Excel save
            print(f"Advanced formatting failed, using simple save: {e}")
            self.extracted_df.to_excel(file_path, index=False)

    def _create_formatted_excel_xlsxwriter(self, file_path):
        """Create formatted Excel using xlsxwriter, writing data manually."""
        try:
            import xlsxwriter
        except ImportError:
            raise ImportError("xlsxwriter not available")
        
        with pd.ExcelWriter(file_path, engine='xlsxwriter') as writer:
            workbook = writer.book
            worksheet = workbook.add_worksheet('Extracted Data')

            # --- Define Formats ---
            header_format = workbook.add_format({'bold': True, 'text_wrap': True, 'valign': 'top', 'fg_color': '#2c3e50', 'font_color': 'white', 'border': 1, 'align': 'center', 'font_size': 12})
            data_format = workbook.add_format({'text_wrap': True, 'valign': 'top', 'border': 1, 'align': 'left', 'font_size': 10})
            number_format = workbook.add_format({'text_wrap': True, 'valign': 'top', 'border': 1, 'align': 'right', 'font_size': 10, 'num_format': '#,##0.00'})
            
            # Add new highlight formats with dark blue background and white text
            highlight_data_format = workbook.add_format({'bg_color': '#4F81BD', 'font_color': '#FFFFFF', 'text_wrap': True, 'valign': 'top', 'border': 1, 'align': 'left', 'font_size': 10})
            highlight_number_format = workbook.add_format({'bg_color': '#4F81BD', 'font_color': '#FFFFFF', 'text_wrap': True, 'valign': 'top', 'border': 1, 'align': 'right', 'font_size': 10, 'num_format': '#,##0.00'})

            # --- Write Header ---
            # Move 'SWITCH_DETAILS_FOLIO_NO' to first column if present and rename to FOLIO_NO
            columns = list(self.extracted_df.columns)
            if 'SWITCH_DETAILS_FOLIO_NO' in columns:
                columns.insert(0, columns.pop(columns.index('SWITCH_DETAILS_FOLIO_NO')))
            columns = ['FOLIO_NO' if col == 'SWITCH_DETAILS_FOLIO_NO' else col for col in columns]
            for col_num, value in enumerate(columns):
                worksheet.write(0, col_num, value, header_format)
            
            # --- Write Data Manually ---
            numeric_cols = ['TRADES_AMOUNT', 'switch in TRAIL_1ST_YEAR', 'switch out TRAIL_1ST_YEAR', 'switch in TRAIL_1ST_YEAR -Previous']
            # Add a yellow highlight format for cell-level highlighting
            highlight_cell_format = workbook.add_format({'bg_color': '#FFD700', 'font_color': '#000000', 'border': 1, 'align': 'left', 'font_size': 10})
            for row_idx, row in self.extracted_df.iterrows():
                excel_row_num = row_idx + 1
                is_highlighted = row_idx in self.rows_to_highlight
                current_data_format = highlight_data_format if is_highlighted else data_format
                current_number_format = highlight_number_format if is_highlighted else number_format
                # Reorder row for folio no first
                row_values = list(row)
                if 'SWITCH_DETAILS_FOLIO_NO' in self.extracted_df.columns:
                    idx = self.extracted_df.columns.get_loc('SWITCH_DETAILS_FOLIO_NO')
                    row_values.insert(0, row_values.pop(idx))
                for col_idx, value in enumerate(row_values):
                    col_name = columns[col_idx]
                    # Highlight RATECATEGORY and RATECATEGORY -Previous if they differ
                    if col_name in ['RATECATEGORY', 'RATECATEGORY -Previous']:
                        ratecat = row.get('RATECATEGORY')
                        ratecat_prev = row.get('RATECATEGORY -Previous')
                        if ratecat != ratecat_prev:
                            if pd.isna(value):
                                worksheet.write_blank(excel_row_num, col_idx, None, highlight_cell_format)
                            elif col_name in numeric_cols:
                                try:
                                    worksheet.write_number(excel_row_num, col_idx, float(value), highlight_cell_format)
                                except (ValueError, TypeError):
                                    worksheet.write_string(excel_row_num, col_idx, str(value), highlight_cell_format)
                            else:
                                worksheet.write_string(excel_row_num, col_idx, str(value), highlight_cell_format)
                            continue
                    if pd.isna(value):
                        worksheet.write_blank(excel_row_num, col_idx, None, current_data_format)
                        continue
                    if col_name in numeric_cols:
                        try:
                            worksheet.write_number(excel_row_num, col_idx, float(value), current_number_format)
                        except (ValueError, TypeError):
                            worksheet.write_string(excel_row_num, col_idx, str(value), current_data_format)
                    else:
                        worksheet.write_string(excel_row_num, col_idx, str(value), current_data_format)

            # --- Set Column Widths ---
            for col_num, col_name in enumerate(columns):
                max_width = len(str(col_name))
                for value in self.extracted_df[col_name if col_name != 'FOLIO_NO' else 'SWITCH_DETAILS_FOLIO_NO'].dropna():
                    max_width = max(max_width, len(str(value)))
                worksheet.set_column(col_num, col_num, min(max_width + 2, 40))

            # --- Add Analytics Sheet ---
            payout_cols = [col for col in self.extracted_df.columns if col.startswith('PAYOUT')]
            # Add scheme type columns
            scheme_type_cols = []
            if 'Scheme Type Swith IN' in self.extracted_df.columns:
                scheme_type_cols.append('Scheme Type Swith IN')
            if 'Scheme Type Swith Out' in self.extracted_df.columns:
                scheme_type_cols.append('Scheme Type Swith Out')
            if 'SWITCH_DETAILS_FOLIO_NO' in self.extracted_df.columns:
                analytics_df = self.extracted_df[self.extracted_df['previous < current switch in TRAIL_1ST_YEAR'] == 'check']
                agg_dict = {
                    'TRADES_AMOUNT': 'sum',
                    'switch in TRAIL_1ST_YEAR': 'first',
                    'switch in TRAIL_1ST_YEAR -Previous': 'first',
                }
                for col in payout_cols:
                    agg_dict[col] = 'sum'
                for col in scheme_type_cols:
                    agg_dict[col] = 'first'
                analytics_summary = analytics_df.groupby(
                    ['SWITCH_DETAILS_FOLIO_NO', 'TRADES_BROK_DLR_CODE', 'switch in'], as_index=False
                ).agg(agg_dict)
                analytics_summary = analytics_summary.rename(columns={
                    'SWITCH_DETAILS_FOLIO_NO': 'FOLIO_NO',
                    'TRADES_BROK_DLR_CODE': 'ARN',
                    'switch in': 'Switch In Scheme Name',
                    'TRADES_AMOUNT': 'TRADES_AMOUNT(sum)',
                    'switch in TRAIL_1ST_YEAR': 'switch in TRAIL_1ST_YEAR (CURRENT)',
                    'switch in TRAIL_1ST_YEAR -Previous': 'switch in TRAIL_1ST_YEAR -Previous'
                })
                analytics_columns = list(analytics_summary.columns)
            else:
                analytics_df = self.extracted_df[self.extracted_df['previous < current switch in TRAIL_1ST_YEAR'] == 'check']
                agg_dict = {
                    'TRADES_AMOUNT': 'sum',
                    'switch in TRAIL_1ST_YEAR': 'first',
                    'switch in TRAIL_1ST_YEAR -Previous': 'first',
                }
                for col in payout_cols:
                    agg_dict[col] = 'sum'
                for col in scheme_type_cols:
                    agg_dict[col] = 'first'
                analytics_summary = analytics_df.groupby(
                    ['TRADES_BROK_DLR_CODE', 'switch in'], as_index=False
                ).agg(agg_dict)
                analytics_summary = analytics_summary.rename(columns={
                    'TRADES_BROK_DLR_CODE': 'ARN',
                    'switch in': 'Switch In Scheme Name',
                    'TRADES_AMOUNT': 'TRADES_AMOUNT(sum)',
                    'switch in TRAIL_1ST_YEAR': 'switch in TRAIL_1ST_YEAR (CURRENT)',
                    'switch in TRAIL_1ST_YEAR -Previous': 'switch in TRAIL_1ST_YEAR -Previous'
                })
                analytics_columns = list(analytics_summary.columns)
            analytics_summary.to_excel(writer, sheet_name='Analytics', index=False, header=False, startrow=2)
            worksheet_analytics = writer.sheets['Analytics']
            num_cols = len(analytics_columns)
            # Title format
            analytics_title_format = workbook.add_format({'bold': True, 'font_size': 20, 'align': 'center', 'valign': 'vcenter', 'fg_color': '#4472C4', 'font_color': 'white'})
            # Sub-header format
            analytics_subheader_format = workbook.add_format({'bold': True, 'font_size': 12, 'align': 'center', 'valign': 'vcenter', 'fg_color': '#D9E1F2', 'font_color': '#4472C4'})
            # Header format
            analytics_header_format = workbook.add_format({'bold': True, 'font_size': 11, 'align': 'center', 'valign': 'vcenter', 'fg_color': '#2F3E5C', 'font_color': 'white', 'border': 1})
            # Data format
            analytics_data_format = workbook.add_format({'font_size': 10, 'border': 1})
            # Title row (row 0)
            worksheet_analytics.merge_range(0, 0, 0, num_cols-1, 'Distributor wise Analytics', analytics_title_format)
            # Sub-header row (row 1)
            worksheet_analytics.merge_range(1, 0, 1, num_cols-1, 'previous < current switch in TRAIL_1ST_YEAR', analytics_subheader_format)
            # Column headers (row 2)
            for col_num, value in enumerate(analytics_columns):
                worksheet_analytics.write(2, col_num, value, analytics_header_format)
            # Data rows (start at row 3)
            for row_idx, row in analytics_summary.iterrows():
                for col_idx, col_name in enumerate(analytics_columns):
                    worksheet_analytics.write(row_idx + 3, col_idx, row[col_name], analytics_data_format)
            # Set column widths
            for col_num, col_name in enumerate(analytics_columns):
                max_width = len(str(col_name))
                for value in analytics_summary[col_name].dropna():
                    max_width = max(max_width, len(str(value)))
                worksheet_analytics.set_column(col_num, col_num, min(max_width + 2, 40))

            # --- Add Switching Rate Analytics Sheet ---
            switching_df = self.extracted_df[self.extracted_df['switching rate check'] == 'check']
            if 'SWITCH_DETAILS_FOLIO_NO' in self.extracted_df.columns:
                agg_dict = {
                    'TRADES_AMOUNT': 'sum',
                    'switch in TRAIL_1ST_YEAR': 'first',
                    'switch out TRAIL_1ST_YEAR': 'first',
                    'switching rate check': 'first',
                }
                for col in payout_cols:
                    agg_dict[col] = 'sum'
                for col in scheme_type_cols:
                    agg_dict[col] = 'first'
                switching_summary = switching_df.groupby(
                    ['SWITCH_DETAILS_FOLIO_NO', 'TRADES_BROK_DLR_CODE', 'switch in', 'switch out'], as_index=False
                ).agg(agg_dict)
                switching_summary = switching_summary.rename(columns={
                    'SWITCH_DETAILS_FOLIO_NO': 'FOLIO_NO',
                    'TRADES_BROK_DLR_CODE': 'ARN',
                    'switch in': 'Switch In Scheme Name',
                    'switch out': 'Switch OUT Scheme Name',
                    'TRADES_AMOUNT': 'TRADES_AMOUNT(sum)',
                    'switch in TRAIL_1ST_YEAR': 'switch in TRAIL_1ST_YEAR',
                    'switch out TRAIL_1ST_YEAR': 'switch OUT TRAIL_1ST_YEAR',
                    'switching rate check': 'switching rate check'
                })
                switching_columns = list(switching_summary.columns)
                switching_summary.to_excel(writer, sheet_name='Switching Rate Analytics', index=False, header=False, startrow=2)
                worksheet_switching = writer.sheets['Switching Rate Analytics']
                num_cols = len(switching_columns)
                worksheet_switching.merge_range(0, 0, 0, num_cols-1, 'Switching Rate Analytics', analytics_title_format)
                worksheet_switching.merge_range(1, 0, 1, num_cols-1, 'switch in TRAIL_1ST_YEAR > switch out TRAIL_1ST_YEAR', analytics_subheader_format)
                for col_num, value in enumerate(switching_columns):
                    worksheet_switching.write(2, col_num, value, analytics_header_format)
                for row_idx, row in switching_summary.iterrows():
                    for col_idx, col_name in enumerate(switching_columns):
                        worksheet_switching.write(row_idx + 3, col_idx, row[col_name], analytics_data_format)
                for col_num, col_name in enumerate(switching_columns):
                    max_width = len(str(col_name))
                    for value in switching_summary[col_name].dropna():
                        max_width = max(max_width, len(str(value)))
                    worksheet_switching.set_column(col_num, col_num, min(max_width + 2, 40))

            # --- Add Direct to Regular Analytics Sheet ---
            direct_df = self.extracted_df[self.extracted_df['Direct to Regular'] == 'check']
            if 'SWITCH_DETAILS_FOLIO_NO' in self.extracted_df.columns:
                agg_dict = {
                    'TRADES_AMOUNT': 'sum',
                }
                for col in payout_cols:
                    agg_dict[col] = 'sum'
                for col in scheme_type_cols:
                    agg_dict[col] = 'first'
                direct_summary = direct_df.groupby(
                    ['SWITCH_DETAILS_FOLIO_NO', 'TRADES_BROK_DLR_CODE', 'switch in', 'switch out'], as_index=False
                ).agg(agg_dict)
                direct_summary = direct_summary.rename(columns={
                    'SWITCH_DETAILS_FOLIO_NO': 'FOLIO_NO',
                    'TRADES_BROK_DLR_CODE': 'ARN',
                    'switch in': 'Switch In Scheme Name',
                    'switch out': 'Switch OUT Scheme Name',
                    'TRADES_AMOUNT': 'TRADES_AMOUNT(sum)'
                })
                direct_columns = list(direct_summary.columns)
                direct_summary.to_excel(writer, sheet_name='Direct to Regular Analytics', index=False, header=False, startrow=2)
                worksheet_direct = writer.sheets['Direct to Regular Analytics']
                num_cols = len(direct_columns)
                worksheet_direct.merge_range(0, 0, 0, num_cols-1, 'Distributor wise Analytics', analytics_title_format)
                worksheet_direct.merge_range(1, 0, 1, num_cols-1, 'Direct to Regular', analytics_subheader_format)
                for col_num, value in enumerate(direct_columns):
                    worksheet_direct.write(2, col_num, value, analytics_header_format)
                for row_idx, row in direct_summary.iterrows():
                    for col_idx, col_name in enumerate(direct_columns):
                        worksheet_direct.write(row_idx + 3, col_idx, row[col_name], analytics_data_format)
                for col_num, col_name in enumerate(direct_columns):
                    max_width = len(str(col_name))
                    for value in direct_summary[col_name].dropna():
                        max_width = max(max_width, len(str(value)))
                    worksheet_direct.set_column(col_num, col_num, min(max_width + 2, 40))

    def _create_formatted_excel_openpyxl(self, file_path):
        """Create formatted Excel using openpyxl as fallback, writing manually."""
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

        wb = Workbook()
        ws = wb.active
        ws.title = 'Extracted Data'

        # --- Define Styles ---
        header_font = Font(bold=True, color="FFFFFF", size=12)
        header_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="top", wrap_text=True)
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        # Add new highlight styles
        highlight_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        highlight_font = Font(size=10, color="FFFFFF")
        
        # --- Write Header ---
        # Move 'SWITCH_DETAILS_FOLIO_NO' to first column if present and rename to FOLIO_NO
        columns = list(self.extracted_df.columns)
        if 'SWITCH_DETAILS_FOLIO_NO' in columns:
            columns.insert(0, columns.pop(columns.index('SWITCH_DETAILS_FOLIO_NO')))
        columns = ['FOLIO_NO' if col == 'SWITCH_DETAILS_FOLIO_NO' else col for col in columns]
        for col_idx, col_name in enumerate(columns, 1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        # --- Write Data ---
        data_font = Font(size=10)
        data_alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        numeric_cols = ['TRADES_AMOUNT', 'switch in TRAIL_1ST_YEAR', 'switch out TRAIL_1ST_YEAR', 'switch in TRAIL_1ST_YEAR -Previous']
        # Add a yellow highlight fill for cell-level highlighting
        highlight_cell_fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
        for row_idx, row in self.extracted_df.iterrows():
            excel_row_num = row_idx + 2  # +2 for 1-based index and header
            is_highlighted = row_idx in self.rows_to_highlight
            row_values = list(row)
            if 'SWITCH_DETAILS_FOLIO_NO' in self.extracted_df.columns:
                idx = self.extracted_df.columns.get_loc('SWITCH_DETAILS_FOLIO_NO')
                row_values.insert(0, row_values.pop(idx))
            for col_idx, value in enumerate(row_values, 1):
                col_name = columns[col_idx - 1]
                if pd.isna(value):
                    value = '' # Replace NaN with blank string
                cell = ws.cell(row=excel_row_num, column=col_idx, value=value)
                cell.alignment = data_alignment
                cell.border = thin_border
                # Highlight RATECATEGORY and RATECATEGORY -Previous if they differ
                if col_name in ['RATECATEGORY', 'RATECATEGORY -Previous']:
                    ratecat = row.get('RATECATEGORY')
                    ratecat_prev = row.get('RATECATEGORY -Previous')
                    if ratecat != ratecat_prev:
                        cell.fill = highlight_cell_fill
                if is_highlighted:
                    cell.font = highlight_font
                    cell.fill = highlight_fill if cell.fill == PatternFill() else cell.fill
                else:
                    cell.font = data_font
                # Set number format for numeric columns
                if col_name in numeric_cols:
                    cell.number_format = '#,##0.00'
        
        # --- Set Column Widths ---
        for col_idx, col_name in enumerate(columns, 1):
            max_width = len(str(col_name))
            for value in self.extracted_df[col_name if col_name != 'FOLIO_NO' else 'SWITCH_DETAILS_FOLIO_NO'].dropna():
                max_width = max(max_width, len(str(value)))
            ws.column_dimensions[chr(ord('A') + col_idx - 1)].width = min(max_width + 2, 40)
            
        # --- Add Analytics Sheet (openpyxl) ---
        payout_cols = [col for col in self.extracted_df.columns if col.startswith('PAYOUT')]
        scheme_type_cols = []
        if 'Scheme Type Swith IN' in self.extracted_df.columns:
            scheme_type_cols.append('Scheme Type Swith IN')
        if 'Scheme Type Swith Out' in self.extracted_df.columns:
            scheme_type_cols.append('Scheme Type Swith Out')
        if 'SWITCH_DETAILS_FOLIO_NO' in self.extracted_df.columns:
            analytics_df = self.extracted_df[self.extracted_df['previous < current switch in TRAIL_1ST_YEAR'] == 'check']
            agg_dict = {
                'TRADES_AMOUNT': 'sum',
                'switch in TRAIL_1ST_YEAR': 'first',
                'switch in TRAIL_1ST_YEAR -Previous': 'first',
            }
            for col in payout_cols:
                agg_dict[col] = 'sum'
            for col in scheme_type_cols:
                agg_dict[col] = 'first'
            analytics_summary = analytics_df.groupby(
                ['SWITCH_DETAILS_FOLIO_NO', 'TRADES_BROK_DLR_CODE', 'switch in'], as_index=False
            ).agg(agg_dict)
            analytics_summary = analytics_summary.rename(columns={
                'SWITCH_DETAILS_FOLIO_NO': 'FOLIO_NO',
                'TRADES_BROK_DLR_CODE': 'ARN',
                'switch in': 'Switch In Scheme Name',
                'TRADES_AMOUNT': 'TRADES_AMOUNT(sum)',
                'switch in TRAIL_1ST_YEAR': 'switch in TRAIL_1ST_YEAR (CURRENT)',
                'switch in TRAIL_1ST_YEAR -Previous': 'switch in TRAIL_1ST_YEAR -Previous'
            })
            analytics_columns = list(analytics_summary.columns)
        else:
            analytics_df = self.extracted_df[self.extracted_df['previous < current switch in TRAIL_1ST_YEAR'] == 'check']
            agg_dict = {
                'TRADES_AMOUNT': 'sum',
                'switch in TRAIL_1ST_YEAR': 'first',
                'switch in TRAIL_1ST_YEAR -Previous': 'first',
            }
            for col in payout_cols:
                agg_dict[col] = 'sum'
            for col in scheme_type_cols:
                agg_dict[col] = 'first'
            analytics_summary = analytics_df.groupby(
                ['TRADES_BROK_DLR_CODE', 'switch in'], as_index=False
            ).agg(agg_dict)
            analytics_summary = analytics_summary.rename(columns={
                'TRADES_BROK_DLR_CODE': 'ARN',
                'switch in': 'Switch In Scheme Name',
                'TRADES_AMOUNT': 'TRADES_AMOUNT(sum)',
                'switch in TRAIL_1ST_YEAR': 'switch in TRAIL_1ST_YEAR (CURRENT)',
                'switch in TRAIL_1ST_YEAR -Previous': 'switch in TRAIL_1ST_YEAR -Previous'
            })  
            analytics_columns = list(analytics_summary.columns)
        analytics_summary = analytics_summary.sort_values(by='ARN').reset_index(drop=True)
        ws_analytics = wb.create_sheet('Analytics')
        num_cols = len(analytics_columns)
        # Title
        ws_analytics.merge_cells(start_row=1, start_column=1, end_row=1, end_column=num_cols)
        cell = ws_analytics.cell(row=1, column=1)
        cell.value = 'Distributor wise Analytics'
        cell.font = Font(bold=True, size=20, color='FFFFFF')
        cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center')
        # Sub-header
        ws_analytics.merge_cells(start_row=2, start_column=1, end_row=2, end_column=num_cols)
        cell = ws_analytics.cell(row=2, column=1)
        cell.value = 'previous < current switch in TRAIL_1ST_YEAR'
        cell.font = Font(bold=True, size=12, color='4472C4')
        cell.fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center')
        # Column headers
        for col_idx, col_name in enumerate(analytics_columns, 1):
            cell = ws_analytics.cell(row=3, column=col_idx, value=col_name)
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color='2F3E5C', end_color='2F3E5C', fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center')
        # Data rows
        for row_idx, row in analytics_summary.iterrows():
            for col_idx, col_name in enumerate(analytics_columns, 1):
                cell = ws_analytics.cell(row=row_idx + 4, column=col_idx, value=row[col_name])
                cell.font = Font(size=10)
                cell.alignment = Alignment(horizontal='left', vertical='center')
                cell.border = thin_border
        # Set column widths
        for col_idx, col_name in enumerate(analytics_columns, 1):
            max_width = len(str(col_name))
            for value in analytics_summary[col_name].dropna():
                max_width = max(max_width, len(str(value)))
            ws_analytics.column_dimensions[chr(ord('A') + col_idx - 1)].width = min(max_width + 2, 40)

        # --- Add Switching Rate Analytics Sheet (openpyxl) ---
        switching_df = self.extracted_df[self.extracted_df['switching rate check'] == 'check']
        if 'SWITCH_DETAILS_FOLIO_NO' in self.extracted_df.columns:
            agg_dict = {
                'TRADES_AMOUNT': 'sum',
                'switch in TRAIL_1ST_YEAR': 'first',
                'switch out TRAIL_1ST_YEAR': 'first',
                'switching rate check': 'first',
            }
            for col in payout_cols:
                agg_dict[col] = 'sum'
            for col in scheme_type_cols:
                agg_dict[col] = 'first'
            switching_summary = switching_df.groupby(
                ['SWITCH_DETAILS_FOLIO_NO', 'TRADES_BROK_DLR_CODE', 'switch in', 'switch out'], as_index=False
            ).agg(agg_dict)
            switching_summary = switching_summary.rename(columns={
                'SWITCH_DETAILS_FOLIO_NO': 'FOLIO_NO',
                'TRADES_BROK_DLR_CODE': 'ARN',
                'switch in': 'Switch In Scheme Name',
                'switch out': 'Switch OUT Scheme Name',
                'TRADES_AMOUNT': 'TRADES_AMOUNT(sum)',
                'switch in TRAIL_1ST_YEAR': 'switch in TRAIL_1ST_YEAR',
                'switch out TRAIL_1ST_YEAR': 'switch OUT TRAIL_1ST_YEAR',
                'switching rate check': 'switching rate check'
            })
            switching_columns = list(switching_summary.columns)
            ws_switching = wb.create_sheet('Switching Rate Analytics')
            num_cols = len(switching_columns)
            ws_switching.merge_cells(start_row=1, start_column=1, end_row=1, end_column=num_cols)
            cell = ws_switching.cell(row=1, column=1)
            cell.value = 'Switching Rate Analytics'
            cell.font = Font(bold=True, size=20, color='FFFFFF')
            cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center')
            ws_switching.merge_cells(start_row=2, start_column=1, end_row=2, end_column=num_cols)
            cell = ws_switching.cell(row=2, column=1)
            cell.value = 'switch in TRAIL_1ST_YEAR > switch out TRAIL_1ST_YEAR'
            cell.font = Font(bold=True, size=12, color='4472C4')
            cell.fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center')
            for col_idx, col_name in enumerate(switching_columns, 1):
                cell = ws_switching.cell(row=3, column=col_idx, value=col_name)
                cell.font = Font(bold=True, color='FFFFFF')
                cell.fill = PatternFill(start_color='2F3E5C', end_color='2F3E5C', fill_type='solid')
                cell.alignment = Alignment(horizontal='center', vertical='center')
            for row_idx, row in switching_summary.iterrows():
                for col_idx, col_name in enumerate(switching_columns, 1):
                    cell = ws_switching.cell(row=row_idx + 4, column=col_idx, value=row[col_name])
                    cell.font = Font(size=10)
                    cell.alignment = Alignment(horizontal='left', vertical='center')
                    cell.border = thin_border
            for col_idx, col_name in enumerate(switching_columns, 1):
                max_width = len(str(col_name))
                for value in switching_summary[col_name].dropna():
                    max_width = max(max_width, len(str(value)))
                ws_switching.column_dimensions[chr(ord('A') + col_idx - 1)].width = min(max_width + 2, 40)

        # --- Add Direct to Regular Analytics Sheet (openpyxl) ---
        direct_df = self.extracted_df[self.extracted_df['Direct to Regular'] == 'check']
        if 'SWITCH_DETAILS_FOLIO_NO' in self.extracted_df.columns:
            agg_dict = {
                'TRADES_AMOUNT': 'sum',
            }
            for col in payout_cols:
                agg_dict[col] = 'sum'
            for col in scheme_type_cols:
                agg_dict[col] = 'first'
            direct_summary = direct_df.groupby(
                ['SWITCH_DETAILS_FOLIO_NO', 'TRADES_BROK_DLR_CODE', 'switch in', 'switch out'], as_index=False
            ).agg(agg_dict)
            direct_summary = direct_summary.rename(columns={
                'SWITCH_DETAILS_FOLIO_NO': 'FOLIO_NO',
                'TRADES_BROK_DLR_CODE': 'ARN',
                'switch in': 'Switch In Scheme Name',
                'switch out': 'Switch OUT Scheme Name',
                'TRADES_AMOUNT': 'TRADES_AMOUNT(sum)'
            })
            direct_columns = list(direct_summary.columns)
            ws_direct = wb.create_sheet('Direct to Regular Analytics')
            num_cols = len(direct_columns)
            ws_direct.merge_cells(start_row=1, start_column=1, end_row=1, end_column=num_cols)
            cell = ws_direct.cell(row=1, column=1)
            cell.value = 'Distributor wise Analytics'
            cell.font = Font(bold=True, size=20, color='FFFFFF')
            cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center')
            ws_direct.merge_cells(start_row=2, start_column=1, end_row=2, end_column=num_cols)
            cell = ws_direct.cell(row=2, column=1)
            cell.value = 'Direct to Regular'
            cell.font = Font(bold=True, size=12, color='4472C4')
            cell.fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center')
            for col_idx, col_name in enumerate(direct_columns, 1):
                cell = ws_direct.cell(row=3, column=col_idx, value=col_name)
                cell.font = Font(bold=True, color='FFFFFF')
                cell.fill = PatternFill(start_color='2F3E5C', end_color='2F3E5C', fill_type='solid')
                cell.alignment = Alignment(horizontal='center', vertical='center')
            for row_idx, row in direct_summary.iterrows():
                for col_idx, col_name in enumerate(direct_columns, 1):
                    cell = ws_direct.cell(row=row_idx + 4, column=col_idx, value=row[col_name])
                    cell.font = Font(size=10)
                    cell.alignment = Alignment(horizontal='left', vertical='center')
                    cell.border = thin_border
            for col_idx, col_name in enumerate(direct_columns, 1):
                max_width = len(str(col_name))
                for value in direct_summary[col_name].dropna():
                    max_width = max(max_width, len(str(value)))
                ws_direct.column_dimensions[chr(ord('A') + col_idx - 1)].width = min(max_width + 2, 40)

    def _add_processing_info_sheet_xlsxwriter(self, workbook):
        """Add a processing information sheet using xlsxwriter"""
        worksheet = workbook.add_worksheet('Processing Information')
        
        # Define formats
        title_format = workbook.add_format({
            'bold': True,
            'font_size': 16,
            'fg_color': '#2c3e50',
            'font_color': 'white',
            'align': 'center',
            'border': 1
        })
        
        header_format = workbook.add_format({
            'bold': True,
            'font_size': 12,
            'fg_color': '#34495e',
            'font_color': 'white',
            'border': 1,
            'align': 'left'
        })
        
        data_format = workbook.add_format({
            'font_size': 11,
            'border': 1,
            'align': 'left' 
        })
        
        # Title
        worksheet.merge_range('A1:B1', 'PROCESSING INFORMATION', title_format)
        
        # Processing details
        info_data = [
            ['Input File', os.path.basename(self.input_file_path) if self.input_file_path else 'Not provided'],
            ['Distributor Files', len(self.distributor_files)],
            ['Previous Month Impalment Files', len(self.impalment_prev_files)],
            ['Processing Date', pd.Timestamp.now().strftime('%Y-%m-%d %H:%M:%S')],
            ['Total Columns Extracted', len(self.extracted_df.columns)],
            ['Total Rows Processed', len(self.extracted_df)]
        ]
        
        # Write processing info
        for row_num, (key, value) in enumerate(info_data):
            worksheet.write(row_num + 3, 0, key, header_format)
            worksheet.write(row_num + 3, 1, value, data_format)
        
        # Add distributor files list if available
        if self.distributor_files:
            worksheet.write(len(info_data) + 5, 0, 'Distributor Files List:', header_format)
            for i, distributor_file in enumerate(self.distributor_files):
                worksheet.write(len(info_data) + 6 + i, 0, f'{i+1}.', data_format)
                worksheet.write(len(info_data) + 6 + i, 1, os.path.basename(distributor_file), data_format)
       
        # Add previous month impalment files list if available
        if self.impalment_prev_files:
            worksheet.write(len(info_data) + 6 + len(self.distributor_files), 0, 'Previous Month Impalment Files List:', header_format)
            for i, prev_file in enumerate(self.impalment_prev_files):
                worksheet.write(len(info_data) + 7 + len(self.distributor_files) + i, 0, f'{i+1}.', data_format)
                worksheet.write(len(info_data) + 7 + len(self.distributor_files) + i, 1, os.path.basename(prev_file), data_format)
        
        # Set column widths
        worksheet.set_column('A:A', 25)
        worksheet.set_column('B:B', 40)

def normalize_fund_name(name):
    if not isinstance(name, str):
        return ''
    # Remove all non-alphanumeric characters and make lowercase
    return re.sub(r'[^a-z0-9]', '', name.lower())

def extract_core_fund_name(name):
    """
    Extracts the core fund name by intelligently removing plan/option suffixes.
    Example: 'Fund Name - Direct Plan Growth' -> 'Fund Name'
    """
    if not isinstance(name, str):
        return ''
    
    # The separator is typically ' - '
    if ' - ' in name:
        parts = name.split(' - ')
        # The last part is a potential plan/option suffix
        last_part = parts[-1].lower()
        
        plan_keywords = ['regular', 'direct', 'growth', 'dividend', 'idcw', 'bonus', 'plan']
        
        # If the last part contains any keyword, assume it's a suffix and remove it
        if any(keyword in last_part for keyword in plan_keywords):
            return ' - '.join(parts[:-1]).strip()
            
    # If no clear suffix is found, return the name as is (but stripped)
    # This prevents incorrectly truncating names that have hyphens for other reasons.
    return name.strip()

def normalize_colname(name):
    return re.sub(r'\s+', '', str(name).lower())  # remove all whitespace and lowercase

def find_header_row(sheet_df, required_cols):
    norm_required = [normalize_colname(col) for col in required_cols]
    for i in range(10):
        row = [str(x).strip() if pd.notna(x) else '' for x in sheet_df.iloc[i]]
        norm_row = [normalize_colname(col) for col in row]
        if all(req in norm_row for req in norm_required):
            return i
    return None

def find_best_sheet(rate_category, brokerage_sheets, threshold=85):
    best_score = 0
    best_key = None
    for sheet_key in brokerage_sheets.keys():
        score = fuzz.ratio(rate_category, sheet_key)
        if score > best_score:
            best_score = score
            best_key = sheet_key
    if best_score >= threshold:
        return best_key
    return None

if __name__ == "__main__":
    root = ctk.CTk()
    app = SwitchExtractorApp(root)
    root.mainloop()
